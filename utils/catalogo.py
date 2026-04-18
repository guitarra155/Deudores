"""
utils/catalogo.py
Catálogo de productos: carga desde Excel o cache JSON, búsqueda por nombre y código.

Optimizaciones para PCs de bajos recursos:
  - Cache JSON: después del primer import, las siguientes cargas usan JSON (~10x más rápido)
  - Nombres pre-procesados en minúsculas al cargar (no en cada búsqueda)
  - Carga en hilo de fondo (no bloquea el arranque de la UI)
  - Índice de códigos de barras con dict O(1)
"""
from __future__ import annotations

import json
import threading
from pathlib import Path
from typing import Callable

try:
    import openpyxl
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False


class Catalogo:
    """Índice en memoria del catálogo de productos."""

    ARCHIVO_XLSX  = "data_storage/catalogo.xlsx"
    ARCHIVO_CACHE = "data_storage/catalogo_cache.json"  # cache ultra-rápido

    def __init__(self, base_dir: Path):
        self._base_dir  = base_dir
        # Estructura optimizada: lista de (nombre_lower, nombre_original, codigo, precio)
        self._productos: list[tuple[str, str, str, float | None]] = []
        # por_codigo: {codigo: (nombre, precio)}
        self._por_codigo: dict[str, tuple[str, float | None]] = {}
        # por_nombre: {nombre: precio}
        self._por_nombre: dict[str, float | None] = {}
        self._cargado   = False
        self._lock      = threading.Lock()

    # ------------------------------------------------------------------ #
    # Carga pública                                                        #
    # ------------------------------------------------------------------ #
    def cargar(self) -> bool:
        """
        Carga sincrónica. Prioridad:
          1. Cache JSON  (rápido ~0.05s para 3500 productos)
          2. Excel xlsx  (lento  ~1-3s,  genera cache al terminar)
        """
        if self._cargar_desde_cache():
            return True
        return self._cargar_desde_xlsx()

    def cargar_en_fondo(self, on_listo: Callable[[], None] | None = None) -> None:
        """
        Carga el catálogo en un hilo separado para no bloquear la UI.
        Llama a on_listo() desde el hilo cuando termina.
        """
        def _run():
            self.cargar()
            if on_listo:
                try:
                    on_listo()
                except Exception:
                    pass

        t = threading.Thread(target=_run, daemon=True)
        t.start()

    def importar_xlsx(self, ruta_origen: str) -> bool:
        """Copia el xlsx a data_storage/, lo parsea y genera el cache JSON."""
        import shutil
        destino = self._base_dir / self.ARCHIVO_XLSX
        try:
            destino.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(ruta_origen, str(destino))
        except Exception:
            return False
        # Resetear estado
        with self._lock:
            self._productos  = []
            self._por_codigo = {}
            self._cargado    = False
        return self._cargar_desde_xlsx()

    # ------------------------------------------------------------------ #
    # Consultas (thread-safe)                                              #
    # ------------------------------------------------------------------ #
    @property
    def disponible(self) -> bool:
        return self._cargado

    @property
    def total(self) -> int:
        return len(self._productos)

    def buscar_nombre(self, texto: str, limite: int = 8) -> list[str]:
        """
        Retorna hasta `limite` nombres que contengan `texto` (case-insensitive).
        Usa nombres pre-procesados en minúsculas → muy rápido.
        """
        if not texto or not self._cargado:
            return []
        q = texto.lower()
        resultados = []
        for nombre_lower, nombre_orig, _, _ in self._productos:
            if q in nombre_lower:
                resultados.append(nombre_orig)
                if len(resultados) >= limite:
                    break
        return resultados

    def buscar_por_codigo(self, codigo: str) -> tuple[str, float | None] | None:
        """Retorna (nombre, precio) para el código de barras, o None. O(1)."""
        if not codigo or not self._cargado:
            return None
        return self._por_codigo.get(codigo.strip())

    def obtener_precio(self, nombre: str) -> float | None:
        """Retorna el precio unitario guardado para un nombre exacto."""
        if not nombre or not self._cargado:
            return None
        return self._por_nombre.get(nombre)

    # ------------------------------------------------------------------ #
    # Internos                                                             #
    # ------------------------------------------------------------------ #
    def _cargar_desde_cache(self) -> bool:
        """Carga desde el JSON cache — ultra rápido."""
        ruta = self._base_dir / self.ARCHIVO_CACHE
        if not ruta.exists():
            return False
        try:
            with open(str(ruta), "r", encoding="utf-8") as f:
                data = json.load(f)
            productos = data.get("productos", [])
            por_codigo = data.get("por_codigo", {})
            if not productos:
                return False
            # Reconstruir estructura optimizada
            lista = []
            por_codigo = {}
            por_nombre = {}
            for p in productos:
                nombre = p["n"]
                precio = p.get("p")
                codigo = p.get("c", "")
                lista.append((nombre.lower(), nombre, codigo, precio))
                if codigo:
                    por_codigo[codigo] = (nombre, precio)
                por_nombre[nombre] = precio

            with self._lock:
                self._productos  = lista
                self._por_codigo = por_codigo
                self._por_nombre = por_nombre
                self._cargado    = True
            return True
        except Exception:
            return False

    def _cargar_desde_xlsx(self) -> bool:
        """Carga desde el xlsx y genera el cache JSON para próximas veces."""
        if not _OPENPYXL_OK:
            return False
        ruta = self._base_dir / self.ARCHIVO_XLSX
        if not ruta.exists():
            return False

        try:
            wb = openpyxl.load_workbook(str(ruta), read_only=True, data_only=True)
        except Exception:
            return False

        # Buscar la hoja con columna PRODUCTO
        rows = None
        for sheet_name in wb.sheetnames:
            ws            = wb[sheet_name]
            candidate     = list(ws.iter_rows(values_only=True))
            if not candidate:
                continue
            header        = [str(c).strip().upper() if c is not None else "" for c in candidate[0]]
            if "PRODUCTO" in header or any("PRODUCT" in h for h in header):
                rows = candidate
                break
        wb.close()

        if rows is None or len(rows) < 2:
            return False

        # Índices de columnas
        header    = [str(c).strip().upper() if c is not None else "" for c in rows[0]]
        idx_prod  = next((i for i, h in enumerate(header) if "PRODUCTO" == h or "PRODUCT" in h), None)
        idx_cod   = next((i for i, h in enumerate(header) if "BARRAS" in h or "BARCODE" in h), None)
        idx_est   = next((i for i, h in enumerate(header) if h in ("ESTADO", "STATUS")), None)
        idx_pre   = next((i for i, h in enumerate(header) if any(kw in h for kw in ("UNITARIO", "PRECIO", "PRICE", "P.UNI"))), None)

        if idx_prod is None:
            return False

        productos: list[tuple[str, str, str, float | None]] = []
        por_codigo: dict[str, tuple[str, float | None]] = {}
        por_nombre: dict[str, float | None] = {}
        _activos = {"ACTIVO", "ACTIVE", "1", "TRUE", "SI", "SÍ"}

        for row in rows[1:]:
            # Filtrar inactivos
            if idx_est is not None and row[idx_est] is not None:
                if str(row[idx_est]).strip().upper() not in _activos:
                    continue

            nombre = str(row[idx_prod]).strip() if row[idx_prod] is not None else ""
            if not nombre or nombre.upper() in ("NONE", "N/A", ""):
                continue

            codigo = ""
            if idx_cod is not None and row[idx_cod] is not None:
                codigo = str(row[idx_cod]).strip()
                if codigo.upper() in ("NONE", "N/A"):
                    codigo = ""

            # Extraer precio
            precio = None
            if idx_pre is not None and row[idx_pre] is not None:
                try:
                    p_val = str(row[idx_pre]).replace("$", "").replace(",", "").strip()
                    precio = float(p_val)
                except Exception:
                    precio = None

            productos.append((nombre.lower(), nombre, codigo, precio))
            if codigo:
                por_codigo[codigo] = (nombre, precio)
            por_nombre[nombre] = precio

        if not productos:
            return False

        with self._lock:
            self._productos  = productos
            self._por_codigo = por_codigo
            self._por_nombre = por_nombre
            self._cargado    = True

        # Guardar cache JSON para próximas cargas (mucho más rápido)
        self._guardar_cache(productos)
        return True

    def _guardar_cache(
        self,
        productos: list[tuple[str, str, str, float | None]],
    ) -> None:
        """Persiste el catálogo en JSON compacto para arranque rápido."""
        ruta = self._base_dir / self.ARCHIVO_CACHE
        try:
            data = {
                "productos":  [{"n": orig, "c": cod, "p": pre} for _, orig, cod, pre in productos],
            }
            with open(str(ruta), "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
        except Exception:
            pass  # La cache es opcional — no es crítico si falla
